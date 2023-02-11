from PySide2.QtWidgets import *
from PySide2.QtUiTools import QUiLoader
from PySide2.QtGui import *
import FDMundergroundwater.onedimensionflow as fo
import FDMundergroundwater.twodimensionsflow as ft
import time
import numpy as np
import threading
import openpyxl


class One_dimension_confined_aquifer_stable_flow(QMainWindow):
    def __init__(self):
        super().__init__()
        # 从文件中加载ui格式
        self.ui = QUiLoader().load("ui/odcasf.ui")
        self.ui.setWindowIcon(QIcon("water.ico"))
        # 监测按钮《计算并绘图》
        self.ui.draw.clicked.connect(self.flow_draw)
        # 监测按钮《返回上一级》
        self.ui.back.clicked.connect(self.return_main)
        # 获取自编库中的类的用法
        self.flow = fo.Confined_aquifer_SF()

    def flow_draw(self):
        self.flow.transmissivity(self.ui.transmissivity.toPlainText())
        self.flow.leakage_recharge(self.ui.leakage_recharge.toPlainText())
        self.flow.l_boundary(self.ui.l_boundary.toPlainText())
        self.flow.r_boundary(self.ui.r_boundary.toPlainText())
        self.flow.step_length(self.ui.step_length.toPlainText())
        self.flow.length(self.ui.length.toPlainText())
        self.flow.draw(self.flow.solve())

    def return_main(self):
        self.ui.close()


class One_dimension_confined_aquifer_unstable_flow(QMainWindow):
    def __init__(self):
        super().__init__()
        # 从文件中加载ui格式
        self.ui = QUiLoader().load("ui/odcausf.ui")
        self.ui.setWindowIcon(QIcon("water.ico"))
        # 数值解存放
        self.solve_fdm = None
        # 解析解存放
        self.solve_as = None
        # 傅里叶级数存放
        self.fourier_series = None
        # 解析解分配CPU核心数
        self.cpu_cores = None
        # 误差存放
        self.error = None
        # 相对误差存放
        self.relative_error = None
        # 空间相对差分步长
        self.relative_step_length = None
        # 时间相对差分步长
        self.relative_step_time = None
        # 压力扩散系数
        self.pressure_diffusion_coefficient = None
        # 监测按钮《计算数值解》
        self.ui.solve.clicked.connect(self.solve)
        # 监测按钮《计算解析解》
        self.ui.solve_analytic_solution.clicked.connect(self.solve_analytic_solution)
        # 监测按钮《绘制数值解》
        self.ui.draw_solve.clicked.connect(self.draw_solve)
        # 监测按钮《绘制解析解》
        self.ui.draw_solve_analytic_solution.clicked.connect(self.draw_solve_analytic_solution)
        # 监测按钮《返回上一级》
        self.ui.back.clicked.connect(self.return_main)
        # 监测按钮《误差对比分析》
        self.ui.error_analysis.clicked.connect(self.error_analysis)
        # 监测按钮《绘制误差表面图》
        self.ui.draw_error.clicked.connect(self.draw_error)
        # 监测按钮《保存日志》
        self.ui.save_date.clicked.connect(self.save_date)
        # 获取自编库中的类的用法
        self.flow = fo.Confined_aquifer_USF()
        # 获取当前系统时间戳
        t = time.localtime()
        # 日志时间
        self.time = str(time.strftime("%Y-%m-%d_%H时%M分%S秒", t))
        self.ui.textBrowser.append('程序运行时间（北京时间）：' + str(time.strftime("%Y-%m-%d %H:%M:%S", t)))
        self.ui.textBrowser.append(self.time)
        wb = openpyxl.Workbook()
        ws1 = wb.create_sheet('info', 0)
        ws1.append(['x轴轴长', '空间差分步长', '相对空间差分步长', 't轴轴长', '时间差分步长', '相对时间差分步长', '左边界', '右边界', '初始条件', '压力扩散系数', '傅里叶级数', '相对误差'])
        wb.save('缓存/' + self.time + 'data.xlsx')

    def solve(self):
        self.flow.l_boundary(self.ui.l_boundary.toPlainText())
        self.flow.r_boundary(self.ui.r_boundary.toPlainText())
        self.flow.step_length(self.ui.step_length.toPlainText())
        self.flow.step_time(self.ui.step_time.toPlainText())
        self.flow.x_length(self.ui.x_length.toPlainText())
        self.flow.t_length(self.ui.t_length.toPlainText())
        self.flow.initial_condition(self.ui.initial_condition.toPlainText())
        # 相对空间差分步长和相对时间差分步长的设定
        self.relative_step_length = self.flow.sl / self.flow.xl
        self.relative_step_time = self.flow.st / self.flow.tl
        # 压力扩散系数的设定
        self.pressure_diffusion_coefficient = self.ui.pressure_diffusion_coefficient.toPlainText()
        # 判断填写的是压力模量系数还是导水系数和贮水系数
        if self.ui.pressure_diffusion_coefficient.toPlainText() == '':
            self.flow.storativity(self.ui.storativity.toPlainText())
            self.flow.transmissivity(self.ui.transmissivity.toPlainText())
            self.flow.leakage_recharge(self.ui.leakage_recharge.toPlainText())
        else:
            self.flow.pressure_diffusion_coefficient(self.ui.pressure_diffusion_coefficient.toPlainText())
            self.flow.transmissivity(1)
            self.flow.leakage_recharge("0")
        # 获取当前系统时间戳
        t = time.localtime()
        self.ui.textBrowser.append(str(time.strftime("%H:%M:%S", t)))
        self.ui.textBrowser.append('正在进行数值解求解')
        self.ui.textBrowser.append(
            '相对空间差分步长：' + str(self.relative_step_length) + '相对时间差分步长：' + str(self.relative_step_time))
        self.ui.textBrowser.append(
            '空间差分步长：' + str(self.flow.sl) + ' 时间差分步长：' + str(self.flow.st))
        self.ui.textBrowser.append('压力扩散系数:' + self.pressure_diffusion_coefficient)
        start_time = time.perf_counter()
        self.solve_fdm = self.flow.solve()
        end_time = time.perf_counter()
        self.ui.textBrowser.append('计算完毕，用时' + str(end_time - start_time) + '秒')

    def draw_solve(self):
        self.flow.draw(self.solve_fdm)

    def solve_analytic_solution_threading(self):
        new_thread = threading.Thread(target=self.solve_analytic_solution())
        new_thread.start()
        new_thread.join()

    def solve_analytic_solution(self):
        self.flow.l_boundary(self.ui.l_boundary.toPlainText())
        self.flow.r_boundary(self.ui.r_boundary.toPlainText())
        self.flow.step_length(self.ui.step_length.toPlainText())
        self.flow.step_time(self.ui.step_time.toPlainText())
        self.flow.x_length(self.ui.x_length.toPlainText())
        self.flow.t_length(self.ui.t_length.toPlainText())
        self.flow.initial_condition(self.ui.initial_condition.toPlainText())
        self.fourier_series = self.ui.spinBox.value()
        self.cpu_cores = self.ui.verticalSlider.value()
        # 相对空间差分步长和相对时间差分步长的设定
        self.relative_step_length = self.flow.sl / self.flow.xl
        self.relative_step_time = self.flow.st / self.flow.tl
        # 压力扩散系数的设定
        self.pressure_diffusion_coefficient = self.ui.pressure_diffusion_coefficient.toPlainText()
        # 判断填写的是压力模量系数还是导水系数和贮水系数
        if self.ui.pressure_diffusion_coefficient.toPlainText() == '':
            self.flow.storativity(self.ui.storativity.toPlainText())
            self.flow.transmissivity(self.ui.transmissivity.toPlainText())
            self.flow.leakage_recharge(self.ui.leakage_recharge.toPlainText())
        else:
            self.flow.pressure_diffusion_coefficient(self.ui.pressure_diffusion_coefficient.toPlainText())
            self.flow.transmissivity(1)
            self.flow.leakage_recharge("0")
        # 获取当前系统时间戳
        t = time.localtime()
        self.ui.textBrowser.append(str(time.strftime("%H:%M:%S", t)))
        self.ui.textBrowser.append('正在进行解析解求解')
        self.ui.textBrowser.append(
            '傅里叶级数解（傅里叶级数取前' + str(self.fourier_series) + '项)，' + '分配CPU核心' + str(
                self.cpu_cores) + '个')
        self.ui.textBrowser.append(
            '相对空间差分步长：' + str(self.relative_step_length) + '相对时间差分步长：' + str(self.relative_step_time))
        self.ui.textBrowser.append(
            '空间差分步长：' + str(self.flow.sl) + ' 时间差分步长：' + str(self.flow.st))
        self.ui.textBrowser.append('压力扩散系数:' + self.pressure_diffusion_coefficient)
        start_time = time.perf_counter()
        self.solve_as = self.flow.solve_multi(fourier_series=self.fourier_series, cpu_cores=self.cpu_cores)
        end_time = time.perf_counter()
        self.ui.textBrowser.append('计算完毕，用时' + str(end_time - start_time) + '秒')

    def draw_solve_analytic_solution(self):
        title = '傅里叶级数解（傅里叶级数取前' + str(self.fourier_series) + '项)，' + '分配CPU核心' + str(
            self.cpu_cores) + '个'
        self.flow.draw(self.solve_as, title=title)

    def return_main(self):
        self.ui.close()

    def error_analysis(self):
        # X轴差分点的数目
        m = int(self.flow.xl / self.flow.sl) + 1
        # 时间轴差分点的数目
        n = int(self.flow.tl / self.flow.st) + 1
        self.error = np.zeros((n, m))
        error_all_abs = 0
        analytic_solution_abs = 0
        for k in range(0, n):  # 对时间进行扫描
            for i in range(0, m):  # 对空间进行扫描
                self.error[k, i] = self.solve_fdm[k, i] - self.solve_as[k, i]
                error_all_abs += abs(self.error[k, i])
                analytic_solution_abs += abs(self.solve_as[k, i])
        self.relative_error = error_all_abs / analytic_solution_abs
        # 获取当前系统时间戳
        t = time.localtime()
        self.ui.textBrowser.append(str(time.strftime("%H:%M:%S", t)))
        self.ui.textBrowser.append('误差分析完毕，相对误差:' + str(self.relative_error * 100) + '%')
        self.ui.textBrowser.append('----------------')
        wb = openpyxl.load_workbook('缓存/' + self.time + 'data.xlsx')
        ws = wb['info']
        ws.append([self.flow.xl, self.flow.sl, self.relative_step_length, self.flow.tl, self.flow.st, self.relative_step_time, self.flow.h_l, self.flow.h_r, self.flow.ic, self.pressure_diffusion_coefficient, self.fourier_series, self.relative_error])
        wb.save('缓存/' + self.time + 'data.xlsx')

    def draw_error(self):
        self.flow.draw(self.error, title='绝对误差表面图')

    def save_date(self):
        # 获取当前系统时间戳
        t = time.localtime()
        filepath = QFileDialog.getExistingDirectory(self.ui, "选择文件存储路径")
        f_ = filepath + '/日志' + str(time.strftime("%Y-%m-%d_%H时%M分%S秒", t)) + '.txt'
        f_xlsx = filepath + '/日志' + str(time.strftime("%Y-%m-%d_%H时%M分%S秒", t)) + '.xlsx'
        file = open(f_, 'w')
        file.write(self.ui.textBrowser.toPlainText())
        file.close()
        wb = openpyxl.load_workbook('缓存/' + self.time + 'data.xlsx')
        wb.save(f_xlsx)
        self.ui.textBrowser.append('日志已保存')


class One_dimension_unconfined_aquifer_stable_flow(QMainWindow):
    def __init__(self):
        super().__init__()
        # 从文件中加载ui格式
        self.ui = QUiLoader().load("ui/oduasf.ui")
        self.ui.setWindowIcon(QIcon("water.ico"))
        # 监测按钮《计算并绘图》
        self.ui.draw.clicked.connect(self.flow_draw)
        # 监测按钮《返回上一级》
        self.ui.back.clicked.connect(self.return_main)
        # 获取自编库中的类的用法
        self.flow = fo.Unconfined_aquifer_SF()

    def flow_draw(self):
        self.flow.hydraulic_conductivity(self.ui.hydraulic_conductivity.toPlainText())
        self.flow.leakage_recharge(self.ui.leakage_recharge.toPlainText())
        self.flow.l_boundary(self.ui.l_boundary.toPlainText())
        self.flow.r_boundary(self.ui.r_boundary.toPlainText())
        self.flow.step_length(self.ui.step_length.toPlainText())
        self.flow.length(self.ui.length.toPlainText())
        self.flow.reference_thickness(self.ui.reference_thickness.toPlainText())
        self.flow.draw(self.flow.solve())

    def return_main(self):
        self.ui.close()


class One_dimension_unconfined_aquifer_unstable_flow(QMainWindow):
    def __init__(self):
        super().__init__()
        # 从文件中加载ui格式
        self.ui = QUiLoader().load("ui/oduausf.ui")
        self.ui.setWindowIcon(QIcon("water.ico"))
        # 监测按钮《计算并绘图》
        self.ui.draw.clicked.connect(self.flow_draw)
        # 监测按钮《返回上一级》
        self.ui.back.clicked.connect(self.return_main)
        # 获取自编库中的类的用法
        self.flow = fo.Unconfined_aquifer_USF()

    def flow_draw(self):
        self.flow.l_boundary(self.ui.l_boundary.toPlainText())
        self.flow.r_boundary(self.ui.r_boundary.toPlainText())
        self.flow.step_length(self.ui.step_length.toPlainText())
        self.flow.x_length(self.ui.x_length.toPlainText())
        self.flow.t_length(self.ui.t_length.toPlainText())
        self.flow.initial_condition(self.ui.initial_condition.toPlainText())
        # 判断填写的是压力模量系数还是渗透系数，给水度和参考厚度
        if self.ui.pressure_diffusion_coefficient.toPlainText() == '':
            self.flow.storativity(self.ui.storativity.toPlainText())
            self.flow.hydraulic_conductivity(self.ui.hydraulic_conductivity.toPlainText())
            self.flow.reference_thickness(self.ui.reference_thickness.toPlainText())
            self.flow.leakage_recharge(self.ui.leakage_recharge.toPlainText())
        else:
            self.flow.pressure_diffusion_coefficient(self.ui.pressure_diffusion_coefficient.toPlainText())
            self.flow.reference_thickness(self.ui.reference_thickness.toPlainText())
            self.flow.hydraulic_conductivity(1)
            self.flow.leakage_recharge("0")
        self.flow.draw(self.flow.solve())

    def return_main(self):
        self.ui.close()


class Two_dimension_confined_aquifer_stable_flow(QMainWindow):
    def __init__(self):
        super().__init__()
        # 从文件中加载ui格式
        self.ui = QUiLoader().load("ui/tdcasf.ui")
        self.ui.setWindowIcon(QIcon("water.ico"))
        # 监测按钮《计算并绘图》
        self.ui.draw.clicked.connect(self.flow_draw)
        # 监测按钮《返回上一级》
        self.ui.back.clicked.connect(self.return_main)
        # 获取自编库中的类的用法
        self.flow = ft.Confined_aquifer_SF()

    def flow_draw(self):
        self.flow.transmissivity(self.ui.transmissivity.toPlainText())
        self.flow.leakage_recharge(self.ui.leakage_recharge.toPlainText())
        self.flow.t_boundary(self.ui.t_boundary.toPlainText())
        self.flow.b_boundary(self.ui.b_boundary.toPlainText())
        self.flow.l_boundary(self.ui.l_boundary.toPlainText())
        self.flow.r_boundary(self.ui.r_boundary.toPlainText())
        self.flow.step_length(self.ui.step_length.toPlainText())
        self.flow.x_length(self.ui.x_length.toPlainText())
        self.flow.y_length(self.ui.y_length.toPlainText())
        self.flow.draw(self.flow.solve())

    def return_main(self):
        self.ui.close()


class Two_dimension_unconfined_aquifer_stable_flow(QMainWindow):
    def __init__(self):
        super().__init__()
        # 从文件中加载ui格式
        self.ui = QUiLoader().load("ui/tduasf.ui")
        self.ui.setWindowIcon(QIcon("water.ico"))
        # 监测按钮《计算并绘图》
        self.ui.draw.clicked.connect(self.flow_draw)
        # 监测按钮《返回上一级》
        self.ui.back.clicked.connect(self.return_main)
        # 获取自编库中的类的用法
        self.flow = ft.Unconfined_aquifer_SF()

    def flow_draw(self):
        self.flow.hydraulic_conductivity(self.ui.hydraulic_conductivity.toPlainText())
        self.flow.leakage_recharge(self.ui.leakage_recharge.toPlainText())
        self.flow.t_boundary(self.ui.t_boundary.toPlainText())
        self.flow.b_boundary(self.ui.b_boundary.toPlainText())
        self.flow.l_boundary(self.ui.l_boundary.toPlainText())
        self.flow.r_boundary(self.ui.r_boundary.toPlainText())
        self.flow.step_length(self.ui.step_length.toPlainText())
        self.flow.x_length(self.ui.x_length.toPlainText())
        self.flow.y_length(self.ui.y_length.toPlainText())
        self.flow.reference_thickness(self.ui.reference_thickness.toPlainText())
        self.flow.draw(self.flow.solve())

    def return_main(self):
        self.ui.close()


class Two_dimension_confined_aquifer_unstable_flow(QMainWindow):
    def __init__(self):
        super().__init__()
        # 从文件中加载ui格式
        self.ui = QUiLoader().load("ui/tdcausf.ui")
        self.ui.setWindowIcon(QIcon("water.ico"))
        # 监测按钮《计算并绘图》
        self.ui.draw.clicked.connect(self.flow_draw)
        # 监测按钮《返回上一级》
        self.ui.back.clicked.connect(self.return_main)
        # 监测按钮《上一时刻》
        self.ui.previous_time.clicked.connect(self.previous_time)
        # 监测按钮《下一时刻》
        self.ui.next_time.clicked.connect(self.next_time)
        # 获取自编库中的类的用法
        self.flow = ft.Confined_aquifer_USF()
        # 存储水头解值的列表
        self.h_all_time = []
        self.time_location = 0  # 时刻位置
        self.time_all = 0  # 抛开初始时刻的所有时刻个数

    def flow_draw(self):
        self.flow.l_boundary(self.ui.l_boundary.toPlainText())
        self.flow.r_boundary(self.ui.r_boundary.toPlainText())
        self.flow.t_boundary(self.ui.t_boundary.toPlainText())
        self.flow.b_boundary(self.ui.b_boundary.toPlainText())
        self.flow.step_length(self.ui.step_length.toPlainText())
        self.flow.step_time(self.ui.step_time.toPlainText())
        self.flow.x_length(self.ui.x_length.toPlainText())
        self.flow.y_length(self.ui.y_length.toPlainText())
        self.flow.t_length(self.ui.t_length.toPlainText())
        self.flow.initial_condition(self.ui.initial_condition.toPlainText())
        self.flow.storativity(self.ui.storativity.toPlainText())
        self.flow.transmissivity(self.ui.transmissivity.toPlainText())
        self.flow.leakage_recharge(self.ui.leakage_recharge.toPlainText())
        self.h_all_time = self.flow.solve()
        self.flow.draw(self.h_all_time[0])  # 绘制初始时刻的水头值
        self.time_all = len(self.h_all_time) - 1
        self.time_location = 0
        self.ui.progressBar.reset()
        self.ui.progressBar.setValue(0)  # 进度条置为0

    def next_time(self):
        self.time_location += 1
        self.flow.draw(self.h_all_time[self.time_location])
        self.ui.progressBar.reset()
        T = int((self.time_location / self.time_all) * 100)
        self.ui.progressBar.setValue(T)  # 设置进度条进度

    def previous_time(self):
        self.time_location -= 1
        self.flow.draw(self.h_all_time[self.time_location])
        self.ui.progressBar.reset()
        T = int((self.time_location / self.time_all) * 100)
        self.ui.progressBar.setValue(T)  # 设置进度条进度

    def return_main(self):
        self.ui.close()


class Two_dimension_unconfined_aquifer_unstable_flow(QMainWindow):
    def __init__(self):
        super().__init__()
        # 从文件中加载ui格式
        self.ui = QUiLoader().load("ui/tduausf.ui")
        self.ui.setWindowIcon(QIcon("water.ico"))
        # 监测按钮《计算并绘图》
        self.ui.draw.clicked.connect(self.flow_draw)
        # 监测按钮《返回上一级》
        self.ui.back.clicked.connect(self.return_main)
        # 监测按钮《上一时刻》
        self.ui.previous_time.clicked.connect(self.previous_time)
        # 监测按钮《下一时刻》
        self.ui.next_time.clicked.connect(self.next_time)
        # 获取自编库中的类的用法
        self.flow = ft.Unconfined_aquifer_USF()
        # 存储水头解值的列表
        self.h_all_time = []
        self.time_location = 0  # 时刻位置
        self.time_all = 0  # 抛开初始时刻的所有时刻个数

    def flow_draw(self):
        self.flow.l_boundary(self.ui.l_boundary.toPlainText())
        self.flow.r_boundary(self.ui.r_boundary.toPlainText())
        self.flow.t_boundary(self.ui.t_boundary.toPlainText())
        self.flow.b_boundary(self.ui.b_boundary.toPlainText())
        self.flow.step_length(self.ui.step_length.toPlainText())
        self.flow.step_time(self.ui.step_time.toPlainText())
        self.flow.x_length(self.ui.x_length.toPlainText())
        self.flow.y_length(self.ui.y_length.toPlainText())
        self.flow.t_length(self.ui.t_length.toPlainText())
        self.flow.initial_condition(self.ui.initial_condition.toPlainText())
        self.flow.storativity(self.ui.storativity.toPlainText())
        self.flow.hydraulic_conductivity(self.ui.hydraulic_conductivity.toPlainText())
        self.flow.reference_thickness(self.ui.reference_thickness.toPlainText())
        self.flow.leakage_recharge(self.ui.leakage_recharge.toPlainText())
        self.h_all_time = self.flow.solve()
        self.flow.draw(self.h_all_time[0])  # 绘制初始时刻的水头值
        self.time_all = len(self.h_all_time) - 1
        self.time_location = 0
        self.ui.progressBar.reset()
        self.ui.progressBar.setValue(0)  # 进度条置为0

    def next_time(self):
        self.time_location += 1
        self.flow.draw(self.h_all_time[self.time_location])
        self.ui.progressBar.reset()
        T = int((self.time_location / self.time_all) * 100)
        self.ui.progressBar.setValue(T)  # 设置进度条进度

    def previous_time(self):
        self.time_location -= 1
        self.flow.draw(self.h_all_time[self.time_location])
        self.ui.progressBar.reset()
        T = int((self.time_location / self.time_all) * 100)
        self.ui.progressBar.setValue(T)  # 设置进度条进度

    def return_main(self):
        self.ui.close()


class Two_dimension_Toth_difficult_baisn(QMainWindow):
    def __init__(self):
        super().__init__()
        # 从文件中加载ui格式
        self.ui = QUiLoader().load("ui/toth.ui")
        self.ui.setWindowIcon(QIcon("water.ico"))
        # 监测按钮《多年平均水位绘图》
        self.ui.draw.clicked.connect(self.flow_draw)
        # 监测按钮《返回上一级》
        self.ui.back.clicked.connect(self.return_main)
        # 获取自编库中的类的用法
        self.flow = ft.Toth_difficult_baisn()

    def flow_draw(self):
        self.flow.basin_length(self.ui.basin_length.toPlainText())
        self.flow.basin_high(self.ui.basin_high.toPlainText())
        self.flow.average_water_level_equation(self.ui.average_water_level_equation.toPlainText())
        self.flow.draw_water_level()

    def return_main(self):
        self.ui.close()
